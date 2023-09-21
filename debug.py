import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
import os

# Load the CSV file into a Pandas DataFrame with header starting from row 4
csv_file = 'LogLookupReport_MR.csv'  # Replace with your file path
df = pd.read_csv(csv_file, header=2)

# Parse the 'DateTime' column with the correct format
df['DateTime'] = pd.to_datetime(df['DateTime'], format='%I:%M %p')

# Define your desired time range for PUTWALL PICKING
start_time_putwall = pd.to_datetime('8:00 PM', format='%I:%M %p')
end_time_putwall = pd.to_datetime('11:59 PM', format='%I:%M %p')


# Define your desired time range for REGULAR PICK
start_time_regular = pd.to_datetime('8:00 PM', format='%I:%M %p')
end_time_regular = pd.to_datetime('11:59 PM', format='%I:%M %p')

# Initialize DataFrames to store pick data
putwall_picking_per_user = pd.DataFrame(columns=['UserID', 'PutwallPickingQuantity'])
regular_pick_per_user = pd.DataFrame(columns=['UserID', 'RegularPickQuantity'])
single_pick_per_user = pd.DataFrame(columns=['UserID', 'SinglePickQuantity'])
replenishment_pick_per_user = pd.DataFrame(columns=['UserID', 'ReplenishmentPickQuantity'])

# Function to modify the 'Action' column based on 'BinLabel' for PUTWALL PICKING
def modify_action_putwall(row):
    action = row['Action']
    bin_label = row['BinLabel']
    
    if action == 'PICKLINE' and bin_label.startswith('MW'):
        return 'PUTWALL PICKING'
    
    return action

# Apply the function to the DataFrame for PUTWALL PICKING
df['Action'] = df.apply(modify_action_putwall, axis=1)

# Filter rows based on the specified time range for PUTWALL PICKING
filtered_df_putwall = df[(df['DateTime'] >= start_time_putwall) & (df['DateTime'] <= end_time_putwall)]

# Count the sum of 'Quantity' for 'PUTWALL PICKING' actions per user within the time range for PUTWALL PICKING
putwall_picking_per_user = filtered_df_putwall[filtered_df_putwall['Action'] == 'PUTWALL PICKING'].groupby('UserID')['Quantity'].sum().reset_index(name='PutwallPickingQuantity')

# Filter rows based on the specified time range for REGULAR PICK
filtered_df_regular = df[(df['DateTime'] >= start_time_regular) & (df['DateTime'] <= end_time_regular)]

# Function to modify the 'Action' column based on 'BinLabel' for REGULAR PICK
def modify_action_regular(row):
    action = row['Action']
    bin_label = row['BinLabel']
    packslip = row['Packslip']
    
    if action == 'PICKLINE': 
        if bin_label.startswith(('1H', '1G', '2E', '2H', '3F', '3H', '23', '2R', '1Y', '1C', '1D', '2D', '3D')) and not packslip.startswith('TR'):
            return 'REGULAR PICK'
    
    return action

# Apply the function to the DataFrame for REGULAR PICK
df['Action'] = df.apply(modify_action_regular, axis=1)

# Count the sum of 'Quantity' for 'REGULAR PICK' actions per user within the time range for REGULAR PICK
regular_pick_per_user = filtered_df_regular[filtered_df_regular['Action'] == 'REGULAR PICK'].groupby('UserID')['Quantity'].sum().reset_index(name='RegularPickQuantity')

# Function to check for "SINGLE PICK"
def is_single_pick(row):
    action = row['Action']
    packslip = row['Packslip']
    datetime = row['DateTime']
    
    if action == 'REPLNISH' and len(str(packslip)) >= 7 and str(packslip)[6] == 'S' and datetime >= start_time_putwall and datetime <= end_time_putwall:
        return True
    
    return False

# Apply the function to the DataFrame to identify "SINGLE PICK"
df['IsSinglePick'] = df.apply(is_single_pick, axis=1)

# Filter rows based on the criteria for "SINGLE PICK"
single_pick_df = df[df['IsSinglePick']]

# Count the sum of 'Quantity' for "SINGLE PICK" actions per user within the time range for PUTWALL PICKING
single_pick_per_user = single_pick_df.groupby('UserID')['Quantity'].sum().reset_index(name='SinglePickQuantity')

# Function to check for "REPLENISHMENT PICK"
def is_replenishment_pick(row):
    action = row['Action']
    packslip = row['Packslip']
    datetime = row['DateTime']
    
    if action == 'REPLNISH' and len(str(packslip)) >= 7 and str(packslip)[6] == 'P' and datetime >= start_time_putwall and datetime <= end_time_putwall:
        return True
    
    return False

# Apply the function to the DataFrame to identify "REPLENISHMENT PICK"
df['IsReplenishmentPick'] = df.apply(is_replenishment_pick, axis=1)

# Filter rows based on the criteria for "REPLENISHMENT PICK"
replenishment_pick_df = df[df['IsReplenishmentPick']]

# Count the sum of 'Quantity' for "REPLENISHMENT PICK" actions per user within the time range for PUTWALL PICKING
replenishment_pick_per_user = replenishment_pick_df.groupby('UserID')['Quantity'].sum().reset_index(name='ReplenishmentPickQuantity')

# Define your desired time range for hourly pick totals
start_time_hourly = pd.to_datetime('8:00 PM', format='%I:%M %p')
end_time_hourly = pd.to_datetime('11:59 PM', format='%I:%M %p')  # Updated end time to 11:59 PM

# Filter rows based on the specified time range for hourly pick totals
filtered_df_hourly = df[(df['DateTime'] >= start_time_hourly) & (df['DateTime'] <= end_time_hourly)]

# Group the filtered DataFrame by hour and sum the 'Quantity' for regular picks
hourly_pick_totals = filtered_df_hourly[filtered_df_hourly['Action'] == 'REGULAR PICK'].groupby(filtered_df_hourly['DateTime'].dt.hour)['Quantity'].sum().reset_index(name='TotalRegularPickQuantity')

# Group the filtered DataFrame by hour and sum the 'Quantity' for Single Picks
single_pick_hourly_totals = single_pick_df.groupby(single_pick_df['DateTime'].dt.hour)['Quantity'].sum().reset_index(name='SinglePickQuantity')

# Group the filtered DataFrame by hour and sum the 'Quantity' for Replenishment Picks
replenishment_hourly_totals = replenishment_pick_df.groupby(replenishment_pick_df['DateTime'].dt.hour)['Quantity'].sum().reset_index(name='ReplenishmentPickQuantity')

# Group the filtered DataFrame by hour and sum the 'Quantity' for Putwall Picks
putwall_hourly_totals = filtered_df_hourly[filtered_df_hourly['Action'] == 'PUTWALL PICKING'].groupby(filtered_df_hourly['DateTime'].dt.hour)['Quantity'].sum().reset_index(name='PutwallPickQuantity')

# Create a new Excel workbook
output_excel_file = 'pick_counts.xlsx'
book = openpyxl.Workbook()

# Write each DataFrame to a different sheet in the Excel file
putwall_picking_sheet = book.create_sheet('PUTWALL PICKING')
regular_pick_sheet = book.create_sheet('REGULAR PICK')
single_pick_sheet = book.create_sheet('SINGLE PICK')
replenishment_pick_sheet = book.create_sheet('REPLENISHMENT PICK')
hourly_pick_totals_sheet = book.create_sheet('Total Units picked by hour')

# Convert DataFrames to lists of lists for writing to Excel
putwall_picking_data = [putwall_picking_per_user.columns.tolist()] + putwall_picking_per_user.values.tolist()
regular_pick_data = filtered_df_regular[filtered_df_regular['Action'] == 'REGULAR PICK']
single_pick_data = [single_pick_per_user.columns.tolist()] + single_pick_per_user.values.tolist()
replenishment_pick_data = [replenishment_pick_per_user.columns.tolist()] + replenishment_pick_per_user.values.tolist()
hourly_pick_totals_data = [hourly_pick_totals.columns.tolist()] + hourly_pick_totals.values.tolist()

# Write the data to Excel sheets
for row_data in putwall_picking_data:
    putwall_picking_sheet.append(row_data)

for row_data in dataframe_to_rows(regular_pick_data, index=False, header=True):
    regular_pick_sheet.append(row_data)

for row_data in single_pick_data:
    single_pick_sheet.append(row_data)

for row_data in replenishment_pick_data:
    replenishment_pick_sheet.append(row_data)

# Add the hourly data to the "Total Units picked by hour" sheet
for row_data in hourly_pick_totals_data:
    hourly_pick_totals_sheet.append(row_data)

# Add the hourly Single Pick data to the "Total Units picked by hour" sheet
single_pick_hourly_totals_data = [single_pick_hourly_totals.columns.tolist()] + single_pick_hourly_totals.values.tolist()
for row_data in single_pick_hourly_totals_data:
    hourly_pick_totals_sheet.append(row_data)

# Add the hourly Replenishment Pick data to the "Total Units picked by hour" sheet
replenishment_hourly_totals_data = [replenishment_hourly_totals.columns.tolist()] + replenishment_hourly_totals.values.tolist()
for row_data in replenishment_hourly_totals_data:
    hourly_pick_totals_sheet.append(row_data)

# Add the hourly Putwall Pick data to the "Total Units picked by hour" sheet
putwall_hourly_totals_data = [putwall_hourly_totals.columns.tolist()] + putwall_hourly_totals.values.tolist()
for row_data in putwall_hourly_totals_data:
    hourly_pick_totals_sheet.append(row_data)

# Calculate the totals for each pick type
total_regular_pick = sum(hourly_pick_totals['TotalRegularPickQuantity'])
total_single_pick = sum(single_pick_hourly_totals['SinglePickQuantity'])
total_replenishment_pick = sum(replenishment_hourly_totals['ReplenishmentPickQuantity'])
total_putwall_pick = sum(putwall_hourly_totals['PutwallPickQuantity'])

# Add the totals to the total row
total_row = ['Total', total_regular_pick, total_single_pick, total_replenishment_pick, total_putwall_pick]
hourly_pick_totals_sheet.append(total_row)

# Save the Excel file
book.save(output_excel_file)

# Check if the file was saved successfully
if os.path.exists(output_excel_file):
    print(f"Excel file '{output_excel_file}' saved successfully.")
else:
    print(f"Failed to save Excel file '{output_excel_file}'.")