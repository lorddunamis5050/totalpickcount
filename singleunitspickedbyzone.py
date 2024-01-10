import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill

def perform_single_pick_byzone(df, book):
    # Define your desired time range for SINGLE PICK
    START_TIME_SINGLE_PICK = pd.to_datetime('8:00 PM', format='%I:%M %p').time()
    END_TIME_SINGLE_PICK = pd.to_datetime('11:59 PM', format='%I:%M %p').time()

    # Ensure 'DateTime' is in datetime format
    df['DateTime'] = pd.to_datetime(df['DateTime'])

    # Assuming START_TIME_PUTWALL and END_TIME_PUTWALL are the same as for SINGLE PICK
    # If they are different, you need to define them separately
    START_TIME_SINGLE_PICK = pd.to_datetime('8:00 PM', format='%I:%M %p')
    END_TIME_SINGLE_PICK = pd.to_datetime('11:59 PM', format='%I:%M %p')

    # Apply the additional time-based filtering
    filtered_df_single = df[(df['DateTime'] >= START_TIME_SINGLE_PICK) & (df['DateTime'] <= END_TIME_SINGLE_PICK)]

    # Extract first two letters from 'BinLabel'
    filtered_df_single['BinGroup'] = filtered_df_single['BinLabel'].str[:2]

    # Function to modify 'Action' based on criteria
    def modify_action_single_pick(row):
        action = row['Action']
        packslip = row['Packslip']
        datetime = row['DateTime']

        # Check if datetime is within the specified range (including date if necessary)
        if ((action == 'REPLNISH' and len(str(packslip)) >= 7 and str(packslip)[6] == 'S' and 
             START_TIME_SINGLE_PICK <= datetime.time() <= END_TIME_SINGLE_PICK) or
            packslip == 'LATE' or packslip == 'LATE-PRIO'):
            return 'SINGLE PICK'
        return action

    # Apply the modification to 'Action'
    filtered_df_single['Action'] = filtered_df_single.apply(modify_action_single_pick, axis=1)

    # Filter for SINGLE PICK actions and get unique 'BinGroup' values
    single_pick_df = filtered_df_single[filtered_df_single['Action'] == 'SINGLE PICK']
    unique_bin_groups = single_pick_df['BinGroup'].unique()

    # Group and sum 'Quantity' for SINGLE PICK actions
    replenishment_counts = single_pick_df.groupby('BinGroup')['Quantity'].sum().abs().to_dict()


    # Create a new sheet in the workbook
    sheet = book.create_sheet("SINGLE UNITS PICKED BY ZONE")

    # Write headings and format them
    sheet.append(['ZONE', 'QUANTITY PICKED'])
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for cell in sheet[1]:
        cell.fill = light_blue_fill

    # Populate the new sheet with filtered data
    for zone in unique_bin_groups:
        qty_picked = replenishment_counts.get(zone, 0)
        sheet.append([zone, qty_picked])

    # Your code to further process the data or save it into 'book' goes here

    return replenishment_counts
