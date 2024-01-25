import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def perform_idle_time_analysis(df, book):
    # Define your desired time range
    start_time = pd.to_datetime('8:00 PM', format='%I:%M %p')
    end_time = pd.to_datetime('11:59 PM', format='%I:%M %p')

    # Filter rows based on the specified time range
    df = df.loc[(df['DateTime'] >= start_time) & (df['DateTime'] <= end_time)].copy()

    # Filter rows to include only specific pick types
    valid_pick_types = ['REPLENISHMENT PICK', 'PUTWALL PICKING', 'REGULAR PICK','SINGLE PICK', 'QUICK MOVE','RESOLVE MOVE','SINGLE PACKING','multi PACKING']
    df = df[df['Action'].isin(valid_pick_types)]



    # Calculate the time difference between each action for each user
    df.loc[:, 'TimeDiff'] = df.sort_values(['UserID', 'DateTime']).groupby('UserID')['DateTime'].diff()

    # For the first action of each user, calculate the time difference from 8:00 PM
    first_action_diff = (df.loc[df.groupby('UserID')['DateTime'].idxmin(), 'DateTime'] - start_time)
    
    # Replace NaT with Timedelta 0
    first_action_diff = first_action_diff.where(first_action_diff > pd.Timedelta(minutes=14), pd.Timedelta(minutes=0))
    
    df.loc[df.groupby('UserID')['DateTime'].idxmin(), 'TimeDiff'] = first_action_diff

    # Convert the 'TimeDiff' column to timedelta64
    df['TimeDiff'] = pd.to_timedelta(df['TimeDiff'])

    # Convert the time difference to minutes
    df.loc[:, 'TimeDiff'] = df['TimeDiff'].dt.total_seconds() / 60

    # Consider a user to be idle if the time difference is more than 14 minutes
    df.loc[:, 'Idle'] = np.where(df['TimeDiff'] > 14, 1, 0)

    # Calculate the total idle time for each user
    idle_time_per_user = df[df['Idle'] == 1].groupby('UserID').agg(
        TotalIdleTime=('TimeDiff', 'sum')
    ).reset_index()

    packing_users = df[df['Action'].isin(['SINGLE PACKING', 'multi PACKING'])]['UserID'].unique()
    idle_time_per_user['Packing'] = idle_time_per_user['UserID'].apply(lambda x: 'PACKING' if x in packing_users else '')

        # Identify users with 'RESOLVE PICKING'
    resolve_picking_users = df[df['Action'] == 'RESOLVE MOVE']['UserID'].unique()
    idle_time_per_user['ResolveMove'] = idle_time_per_user['UserID'].apply(lambda x: 'RESOLVE MOVE' if x in resolve_picking_users else '')


    # Filter out users involved in 'PACKING' or 'RESOLVE MOVE'
    filtered_idle_time = idle_time_per_user[(idle_time_per_user['Packing'] == '') & (idle_time_per_user['ResolveMove'] == '')]

 # Find first action time for each user
    first_action_time = df.groupby('UserID')['DateTime'].min().reset_index()
    first_action_time.rename(columns={'DateTime': 'FirstActionTime'}, inplace=True)
    first_action_time['FirstActionTime'] = first_action_time['FirstActionTime'].dt.strftime('%I:%M:%S %p')

     # Find last action time for each user
    last_action_time = df.groupby('UserID')['DateTime'].max().reset_index()
    last_action_time.rename(columns={'DateTime': 'LastActionTime'}, inplace=True)
    last_action_time['LastActionTime'] = last_action_time['LastActionTime'].dt.strftime('%I:%M:%S %p')

    # Merge the first and last action time with the idle time DataFrame
    final_df = pd.merge(filtered_idle_time, first_action_time, on='UserID', how='left')
    final_df = pd.merge(final_df, last_action_time, on='UserID', how='left')


        # Sort final_df by TotalIdleTime in descending order
    final_df.sort_values('TotalIdleTime', ascending=False, inplace=True)


    



    
    # Create the "IDLE TIME" sheet if it doesn't exist
    if 'IDLE TIME' not in book.sheetnames:
        idle_time_sheet = book.create_sheet('IDLE TIME')
    else:
        idle_time_sheet = book['IDLE TIME']


            # Define styles for headers and data cells
    header_font = Font(bold=True)
    data_font = Font(size=12)
    align_center = Alignment(horizontal='center')
    align_left = Alignment(horizontal='left')  # For left alignment
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    idle_time_fill = PatternFill(start_color="EEF5FF", end_color="EEF5FF", fill_type="solid")  # Highlight color for TotalIdleTime

     # Write and format the header row
    header_row = ['UserID', 'TotalIdleTime', 'FirstActionTime', 'LastActionTime']
    idle_time_sheet.append(header_row)
    for cell in idle_time_sheet["1:1"]:
        cell.font = header_font
        cell.alignment = align_center
        cell.fill = header_fill
  


   # Write and format data to the Excel sheet
    for row in final_df.itertuples():
        idle_time_sheet.append([row.UserID, row.TotalIdleTime, row.FirstActionTime, row.LastActionTime])
        row_id = idle_time_sheet.max_row
        for index, cell in enumerate(idle_time_sheet[row_id]):
            cell.font = data_font
            
            # Apply specific formatting for TotalIdleTime column
            if index ==  1:  # Assuming TotalIdleTime is the second column
                cell.alignment = align_left
                cell.fill = idle_time_fill
            if index == 0:
                cell.alignment = align_left

            else:
                cell.alignment = align_center

                    # Adjust column widths
    for column_cells in idle_time_sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        idle_time_sheet.column_dimensions[column_cells[0].column_letter].width = length

            


    print("IDLE TIME analysis completed.")
