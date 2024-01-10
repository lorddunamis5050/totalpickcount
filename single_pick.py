import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from openpyxl.styles import PatternFill

def perform_single_pick_analysis(df , book):
    # Define your desired time range for PUTWALL PICKING
    START_TIME_PUTWALL = pd.to_datetime('8:00 PM', format='%I:%M %p')
    END_TIME_PUTWALL = pd.to_datetime('11:59 PM', format='%I:%M %p')

    df['DateTime'] = pd.to_datetime(df['DateTime'])



    # Initialize DataFrames to store single pick data
    single_pick_per_user = pd.DataFrame(columns=['UserID', 'SinglePickQuantity'])

    # Function to check for "SINGLE PICK"

    
    def modify_action_single_pick(row):
        action = row['Action']
        packslip = row['Packslip']
        datetime = row['DateTime']

        if (action == 'REPLNISH' and len(str(packslip)) >= 7 and str(packslip)[6] == 'S' and 
            datetime >= START_TIME_PUTWALL and datetime <= END_TIME_PUTWALL) or packslip == 'LATE' or  packslip == 'LATE-PRIO':
         return 'SINGLE PICK'

        return action



    # Apply the function to the DataFrame to identify "SINGLE PICK"
    df['Action'] = df.apply(modify_action_single_pick, axis=1)


    # Filter rows based on the criteria for "SINGLE PICK"
    

    filtered_df_single = df[(df['DateTime'] >= START_TIME_PUTWALL) & (df['DateTime'] <=END_TIME_PUTWALL)]

             # Find the highest time worked by any user within the time range
    highest_hours_worked = (filtered_df_single.groupby('UserID')['DateTime']
                             .agg(lambda x: (x.max() - x.min()).total_seconds() / 3600)
                             .max())


    # Count the sum of 'Quantity' for "SINGLE PICK" actions per user within the time range for PUTWALL PICKING
    single_pick_per_user = filtered_df_single[filtered_df_single['Action'] == 'SINGLE PICK'].groupby('UserID').agg(
        SinglePickQuantity=('Quantity', 'sum')
    ).reset_index()

    def calculate_single_pick_time(group):
        group = group.sort_values('DateTime')
        
        # Calculate the duration in minutes to the next row
        # No need to shift the 'DateTime' column before calculating the difference
        group['Duration'] = group['DateTime'].diff().dt.total_seconds().div(60).abs()

        # Mark the rows that are not part of continuous putwall picking
        # Also consider if the duration to the next action is more than 10 minutes as a gap
        group['IsGap'] = (~group['Action'].eq('SINGLE PICK') |
                        group['Action'].shift().ne('SINGLE PICK') |
                        (group['Duration'] > 10))

        # Cumulatively sum the gap marks to create unique session IDs
        group['SessionId'] = group['IsGap'].cumsum()

        # Filter out non-putwall picking rows and rows that start a new session
        replenishment_sessions = group[group['Action'].eq('SINGLE PICK') & ~group['IsGap']]

        # Calculate start and end time for each session
        session_times = replenishment_sessions.groupby('SessionId').agg({'DateTime': ['min', 'max']}) 

        # Calculate duration for each session
        session_durations = (session_times['DateTime']['max']  - session_times['DateTime']['min']).dt.total_seconds().div(60) 

        return session_durations.sum()  # Return the total time across all sessions
        # Filter for putwall picking actions only


    single_pick_actions = filtered_df_single[filtered_df_single['Action'] == 'SINGLE PICK']

            # Calculate total putwall picking time for each user considering gaps
    total_single_pick_time = single_pick_actions.groupby('UserID').apply(calculate_single_pick_time).reset_index(name='Time')

        # Merge this time with the putwall_picking_per_user DataFrame
    single_pick_per_user = single_pick_per_user.merge(total_single_pick_time, on='UserID', how='left')

    # Calculate UPH (Units Per Hour) for each user using the total putwall picking time
    single_pick_per_user['UPH'] = single_pick_per_user.apply(
    lambda row: (row['SinglePickQuantity'] * 60) / row['Time'] if row['Time'] >= 30 else 0, axis=1
    )

    # Convert both "PutwallPickingQuantity" and "UPH" values to their absolute values
    single_pick_per_user['SinglePickQuantity'] = abs(single_pick_per_user['SinglePickQuantity'])
    single_pick_per_user['UPH'] = abs(single_pick_per_user['UPH']).round(1)

        # Calculate the average UPH, excluding zeros
    average_uph = single_pick_per_user.loc[single_pick_per_user['UPH'] > 0, 'UPH'].mean().round(1)

                # Replace NaN or infinite values with zero
    single_pick_per_user.replace([np.inf, -np.inf, np.nan], 0, inplace=True)

    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")


    # Create the "REPLENISHMENT  PICKING" sheet if it doesn't exist
    if 'SINGLE PICK' not in book.sheetnames:
        single_pick_sheet = book.create_sheet('SINGLE PICK')
    else:
        single_pick_sheet = book['SINGLE PICK']

    # Write the header row
    header_row = ['UserID', 'SinglePickQuantity', 'Time','UPH',]
    single_pick_sheet.append(header_row)

                # Format header row with light blue background
    for cell in single_pick_sheet[1]:
        cell.fill = light_blue_fill

                # Sort by UPH in descending order
    single_pick_per_user.sort_values(by='UPH', ascending=False, inplace=True) 

                # Calculate the average UPH, excluding zeros, and round to 2 decimal places
    average_uph = single_pick_per_user.loc[single_pick_per_user['UPH'] > 0, 'UPH'].mean().round(1)  

    # Convert the DataFrame to a list of lists for writing to Excel
    single_picking_data = single_pick_per_user[['UserID', 'SinglePickQuantity', 'Time','UPH']].values.tolist()


        # Write the data to the Excel sheet
    for row_data in single_picking_data:
        single_pick_sheet.append(row_data)

             # Write the average UPH row to the Excel sheet
    average_uph_row = ["Average UPH", "", "", average_uph]
    single_pick_sheet.append(average_uph_row)  

    for cell in single_pick_sheet[single_pick_sheet.max_row]:
        cell.fill = light_blue_fill 



    print("Effufilment PICKING analysis completed.")

