import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from openpyxl.styles import PatternFill

def perform_replenishment_pick_analysis(df, book):
    # Define your desired time range for PUTWALL PICKING
    START_TIME_REGULAR = pd.to_datetime('8:00 PM', format='%I:%M %p')
    END_TIME_REGULAR = pd.to_datetime('11:59 PM', format='%I:%M %p')

    df['DateTime'] = pd.to_datetime(df['DateTime'])

    # Initialize DataFrames to store replenishment pick data
    replenishment_pick_per_user = pd.DataFrame(columns=['UserID', 'ReplenishmentPickQuantity'])

    def modify_action(row):
        action = row['Action']
        packslip = row['Packslip']

        if action == 'REPLNISH' and len(str(packslip)) >= 7 and str(packslip)[6] == 'P':
            return 'REPLENISHMENT PICK'
        elif packslip == 'BIG' or packslip == 'TR_GROUP2L':
            return 'REPLENISHMENT PICK'  # OR do something else for 'BIG' or 'TR_GROUP2L'

        return action


# Apply the modification function to the 'Action' column
    df['Action'] = df.apply(modify_action, axis=1)

    filtered_df_regular = df[(df['DateTime'] >= START_TIME_REGULAR) & (df['DateTime'] <= END_TIME_REGULAR)]



    # Count the sum of 'Quantity' for "REPLENISHMENT PICK" actions per user within the time range for PUTWALL PICKING
    replenishment_pick_per_user = filtered_df_regular[filtered_df_regular['Action'] == 'REPLENISHMENT PICK'].groupby('UserID').agg(
        ReplenishmentPickQuantity = ('Quantity', 'sum')
    ).reset_index()

    def calculate_replenishment_pick_time(group):
        group = group.sort_values('DateTime')
        
        # Calculate the duration in minutes to the next row
        # No need to shift the 'DateTime' column before calculating the difference
        group['Duration'] = group['DateTime'].diff().dt.total_seconds().div(60).abs()

        # Mark the rows that are not part of continuous putwall picking
        # Also consider if the duration to the next action is more than 10 minutes as a gap
        group['IsGap'] = (~group['Action'].eq('REPLENISHMENT PICK') |
                        group['Action'].shift().ne('REPLENISHMENT PICK') |
                        (group['Duration'] > 10))

        # Cumulatively sum the gap marks to create unique session IDs
        group['SessionId'] = group['IsGap'].cumsum()

        # Filter out non-putwall picking rows and rows that start a new session
        replenishment_sessions = group[group['Action'].eq('REPLENISHMENT PICK') & ~group['IsGap']]

        # Calculate start and end time for each session
        session_times = replenishment_sessions.groupby('SessionId').agg({'DateTime': ['min', 'max']}) 

        # Calculate duration for each session
        session_durations = (session_times['DateTime']['max']  - session_times['DateTime']['min']).dt.total_seconds().div(60) 

        return session_durations.sum()  # Return the total time across all sessions
        # Filter for putwall picking actions only

    replenishment_pick_actions = filtered_df_regular[filtered_df_regular['Action'] == 'REPLENISHMENT PICK']

        # Calculate total putwall picking time for each user considering gaps
    total_replenishment_pick_time = replenishment_pick_actions.groupby('UserID').apply(calculate_replenishment_pick_time).reset_index(name='Time')


        # Merge this time with the putwall_picking_per_user DataFrame
    replenishment_pick_per_user = replenishment_pick_per_user.merge(total_replenishment_pick_time, on='UserID', how='left')


    # Calculate UPH (Units Per Hour) for each user using the total putwall picking time
    replenishment_pick_per_user['UPH'] = replenishment_pick_per_user.apply(
    lambda row: (row['ReplenishmentPickQuantity'] * 60) / row['Time'] if row['Time'] >= 30 else 0, axis=1
    )


            # Convert both "PutwallPickingQuantity" and "UPH" values to their absolute values
    replenishment_pick_per_user['ReplenishmentPickQuantity'] = abs(replenishment_pick_per_user['ReplenishmentPickQuantity'])
    replenishment_pick_per_user['UPH'] = abs(replenishment_pick_per_user['UPH']).round(1)

        # Calculate the average UPH, excluding zeros
    average_uph = replenishment_pick_per_user.loc[replenishment_pick_per_user['UPH'] > 0, 'UPH'].mean().round(1)

            # Replace NaN or infinite values with zero
    replenishment_pick_per_user.replace([np.inf, -np.inf, np.nan], 0, inplace=True)

    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")


            # Create the "REPLENISHMENT  PICKING" sheet if it doesn't exist
    if 'REPLENISHMENT PICK' not in book.sheetnames:
        replenishment_pick_sheet = book.create_sheet('REPLENISHMENT PICK')
    else:
        replenishment_pick_sheet = book['REPLENISHMENT PICK']

                    # Write the header row
    header_row = ['UserID', 'ReplenishmentPickQuantity', 'Time', 'UPH']
    replenishment_pick_sheet.append(header_row)

            # Format header row with light blue background
    for cell in replenishment_pick_sheet[1]:
        cell.fill = light_blue_fill


        # Sort by UPH in descending order
    replenishment_pick_per_user.sort_values(by='UPH', ascending=False, inplace=True)    

            # Calculate the average UPH, excluding zeros, and round to 2 decimal places
    average_uph = replenishment_pick_per_user.loc[replenishment_pick_per_user['UPH'] > 0, 'UPH'].mean().round(1)


            # Convert the DataFrame to a list of lists for writing to Excel
    replenishment_picking_data = replenishment_pick_per_user[['UserID', 'ReplenishmentPickQuantity','Time', 'UPH']].values.tolist()

        # Write the data to the Excel sheet
    for row_data in replenishment_picking_data:
        replenishment_pick_sheet.append(row_data)

         # Write the average UPH row to the Excel sheet
    average_uph_row = ["Average UPH", "", "", average_uph]
    replenishment_pick_sheet.append(average_uph_row)  

    for cell in replenishment_pick_sheet[replenishment_pick_sheet.max_row]:
        cell.fill = light_blue_fill 




    print("ByGroup PICKING analysis completed.")


