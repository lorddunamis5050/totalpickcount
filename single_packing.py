import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from openpyxl.styles import PatternFill

def perform_single_pack_analysis(df, book):
    # Define your desired time range for PUTWALL PICKING
    start_time_putwall = pd.to_datetime('8:00 PM', format='%I:%M %p')
    end_time_putwall = pd.to_datetime('11:59 PM', format='%I:%M %p')


    df['DateTime'] = pd.to_datetime(df['DateTime'])

    # Initialize a DataFrame to store PUTWALL PICKING data per user
    putwall_picking_per_user = pd.DataFrame(columns=['UserID', 'SinglePackingQuantity'])

    # Function to modify the 'Action' column based on 'BinLabel' for PUTWALL PICKING
    def modify_action_single_packing(row):
        action = row['Action']
        bin_label = row['BinLabel']
        
        if action == 'PICKLINE' and bin_label.startswith('SH'):
            return 'SINGLE PACKING'
        
        return action

    # Apply the function to the DataFrame for PUTWALL PICKING
    df['Action'] = df.apply(modify_action_single_packing, axis=1)

  # Filter rows based on the specified time range for PUTWALL PICKING
    filtered_df_putwall = df[(df['DateTime'] >= start_time_putwall) & (df['DateTime'] <= end_time_putwall)]


    # Group by "UserID" and calculate total Units picked
    putwall_picking_per_user = filtered_df_putwall[filtered_df_putwall['Action'] == 'SINGLE PACKING'].groupby('UserID').agg(
        SinglePackingQuantity=('Quantity', 'sum')
    ).reset_index()

    def calculate_putwall_picking_time(group):
        group = group.sort_values('DateTime')
        
        # Calculate the duration in minutes to the next row
        # No need to shift the 'DateTime' column before calculating the difference
        group['Duration'] = group['DateTime'].diff().dt.total_seconds().div(60).abs()

        # Mark the rows that are not part of continuous putwall picking
        # Also consider if the duration to the next action is more than 10 minutes as a gap
        group['IsGap'] = (~group['Action'].eq('SINGLE PACKING') |
                        group['Action'].shift().ne('SINGLE PACKING') |
                        (group['Duration'] > 10))

        # Cumulatively sum the gap marks to create unique session IDs
        group['SessionId'] = group['IsGap'].cumsum()

        # Filter out non-putwall picking rows and rows that start a new session
        putwall_sessions = group[group['Action'].eq('SINGLE PACKING') & ~group['IsGap']]

        # Calculate start and end time for each session
        session_times = putwall_sessions.groupby('SessionId').agg({'DateTime': ['min', 'max']}) 

        # Calculate duration for each session
        session_durations = (session_times['DateTime']['max']  - session_times['DateTime']['min']).dt.total_seconds().div(60) 

        return session_durations.sum()  # Return the total time across all sessions



    # Filter for putwall picking actions only
    putwall_picking_actions = filtered_df_putwall[filtered_df_putwall['Action'] == 'SINGLE PACKING']

    # Calculate total putwall picking time for each user considering gaps
    total_putwall_picking_time = putwall_picking_actions.groupby('UserID').apply(calculate_putwall_picking_time).reset_index(name='Time')

    # Merge this time with the putwall_picking_per_user DataFrame
    putwall_picking_per_user = putwall_picking_per_user.merge(total_putwall_picking_time, on='UserID', how='left')


# Calculate UPH (Units Per Hour) for each user using the total putwall picking time
    putwall_picking_per_user['UPH'] = putwall_picking_per_user.apply(
    lambda row: (row['SinglePackingQuantity'] * 60) / row['Time'] if row['Time'] >= 30 else 0, axis=1
    )

    # Calculate UPH for each user, using the highest time as the denominator
    # putwall_picking_per_user['UPH'] = putwall_picking_per_user['SinglePackingQuantity'] / highest_hours_worked

        # Convert both "SinglePackingQuantity" and "UPH" values to their absolute values
    putwall_picking_per_user['SinglePackingQuantity'] = abs(putwall_picking_per_user['SinglePackingQuantity'])
    putwall_picking_per_user['UPH'] = abs(putwall_picking_per_user['UPH']).round(2)

    # Calculate the average UPH, excluding zeros
    average_uph = putwall_picking_per_user.loc[putwall_picking_per_user['UPH'] > 0, 'UPH'].mean()

        # Replace NaN or infinite values with zero
    putwall_picking_per_user.replace([np.inf, -np.inf, np.nan], 0, inplace=True)

    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")


    # Create the "PUTWALL PICKING" sheet if it doesn't exist
    if 'SINGLE PACKING' not in book.sheetnames:
        putwall_picking_sheet = book.create_sheet('SINGLE PACKING')
    else:
        putwall_picking_sheet = book['SINGLE PACKING']

    # Update the header row to include 'Time'
    header_row = ['UserID', 'SinglePackingQuantity', 'Time', 'UPH']
    putwall_picking_sheet.append(header_row)

        # Format header row with light blue background
    for cell in putwall_picking_sheet[1]:
        cell.fill = light_blue_fill

    # Convert the DataFrame to a list of lists for writing to Excel
        # Sort by UPH in descending order
    putwall_picking_per_user.sort_values(by='UPH', ascending=False, inplace=True)


        # Calculate the average UPH, excluding zeros, and round to 2 decimal places
    if putwall_picking_per_user.loc[putwall_picking_per_user['UPH'] > 0, 'UPH'].empty:
        average_uph = 0  # Or any default value you want to use in case of no data
    else:
        average_uph = putwall_picking_per_user.loc[putwall_picking_per_user['UPH'] > 0, 'UPH'].mean().round(1)


    
    putwall_picking_data = putwall_picking_per_user[['UserID', 'SinglePackingQuantity', 'Time', 'UPH']].values.tolist()


    # Write the data to the Excel sheet
    for row_data in putwall_picking_data:
        putwall_picking_sheet.append(row_data)
        
    # Write the average UPH row to the Excel sheet
    average_uph_row = ["Average UPH", "", "", average_uph]
    putwall_picking_sheet.append(average_uph_row)

    for cell in putwall_picking_sheet[putwall_picking_sheet.max_row]:
        cell.fill = light_blue_fill

    print("SINGLE PACKING analysis completed.")

