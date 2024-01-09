import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Define the time range
START_TIME = pd.to_datetime('8:00 PM', format='%I:%M %p')
END_TIME = pd.to_datetime('11:59 PM', format='%I:%M %p')

def calculate_pack_totals(df, action_filter):
    # Filter the DataFrame by action_filter and time range
    filtered_df = df[(df['Action'].isin(action_filter)) & (df['DateTime'].dt.hour >= 20) & (df['DateTime'].dt.hour < 24)]
    
    # Group by hour and sum quantities
    pack_totals = filtered_df.groupby(filtered_df['DateTime'].dt.hour)['Quantity'].apply(lambda x: abs(x).sum()).to_dict()
    
    return pack_totals

def perform_hourly_pack_totals_analysis(df, book):
    # Define the pack types and corresponding action filters
    pack_types = {
        'Single Pack': ['SINGLE PACKING'],
        'Multi Pack': ['multi PACKING'],
    }

    # Create the "Total Units packed by hour" sheet if it doesn't exist
    if 'Total Units packed by hour' not in book.sheetnames:
        hourly_pack_totals_sheet = book.create_sheet('Total Units packed by hour')
    else:
        hourly_pack_totals_sheet = book['Total Units packed by hour']

    # Write the header row
    header_row = ['Hour'] + list(pack_types.keys())
    hourly_pack_totals_sheet.append(header_row)

    # Calculate and write the total quantity for each hour and pack type
    for hour in range(20, 24):
        hour_data = [hour]

        for pack_type, action_filter in pack_types.items():
            pack_totals = calculate_pack_totals(df, action_filter)
            quantity = pack_totals.get(hour, 0)
            hour_data.append(quantity)

        hourly_pack_totals_sheet.append(hour_data)

    # Calculate and write the total quantities for each pick type
    total_row = ['Total']
    for pack_type, action_filter in pack_types.items():
        pack_totals = calculate_pack_totals(df, action_filter)
        total_quantity = sum(pack_totals.values())
        total_row.append(total_quantity)

    hourly_pack_totals_sheet.append(total_row)

    # Save the Excel file
    output_excel_file = 'pick_counts.xlsx'
    book.save(output_excel_file)

    print("Hourly packs totals analysis completed and saved.")

# Usage:
# Call perform_hourly_pick_totals_analysis with your DataFrame and Excel book as arguments
# Example: perform_hourly_pick_totals_analysis(df, openpyxl.Workbook())
